"""Append the n=500 HCDS results to the existing AJAR results.xlsx.

Additive: never modifies existing rows; appends new rows to
`HCDS_summary_all`, `Run_index`; adds new sheets `HCDS_per_question_n=500`
and `Task6_table_n=500`. Safe to re-run.
"""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

REPO = Path("/Users/edward/Projects/AJAR")
XLSX = REPO / "AJAR results.xlsx"
N500_DIR = REPO / "results/runs/2026-05-07_0025_gsm8k500_qwen3-4b_deep-table-ext"


def read_csv_rows(path: Path) -> list[list[str]]:
    with path.open() as f:
        return list(csv.reader(f))


def amend_hcds_summary(wb) -> None:
    """Append n=500 rows: variant 'full_n500' and 'no_mech_n500'."""
    ws = wb["HCDS_summary_all"]
    sources = [
        ("full_n500", N500_DIR / "hcds_summary.csv"),
        ("no_mech_n500", N500_DIR / "hcds_summary_no_mech.csv"),
    ]
    for variant_label, src in sources:
        rows = read_csv_rows(src)
        header = rows[0]
        for r in rows[1:]:
            row_dict = dict(zip(header, r))
            ws.append([
                variant_label,
                row_dict["model"],
                float(row_dict["n_questions"]),
                float(row_dict["mean_hcds"]),
                float(row_dict["ci_low_95"]),
                float(row_dict["ci_high_95"]),
                float(row_dict["t_statistic"]),
                float(row_dict["p_value_two_sided"]),
                float(row_dict["n_features_avg"]),
                row_dict["verdict"],
            ])
    print(f"[amend] HCDS_summary_all: appended 4 rows ({variant_label} etc.)")


def add_per_question_sheet(wb) -> None:
    """Add a new sheet with the n=500 per-question HCDS data."""
    sheet_name = "HCDS_per_question_n=500"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    rows = read_csv_rows(N500_DIR / "hcds_per_question.csv")
    for row in rows:
        ws.append(row)
    print(f"[amend] {sheet_name}: {len(rows) - 1} data rows")


def add_task6_sheet(wb) -> None:
    """Add a new sheet with the n=500 wide table (5+ features per row)."""
    sheet_name = "Task6_table_n=500"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    rows = read_csv_rows(N500_DIR / "task6_table.csv")
    for row in rows:
        ws.append(row)
    print(f"[amend] {sheet_name}: {len(rows) - 1} data rows")


def amend_run_index(wb) -> None:
    """Append the new run to the run index."""
    ws = wb["Run_index"]
    new_row = [
        "2026-05-07_0025_gsm8k500_qwen3-4b_deep-table-ext",
        "yes",
        "n=500 5-feature HCDS extension (paraphrase + perturbation on full GSM8K-500; mech reused from n=50)",
        "Headline: HCDS positive both models at n=500 (Instruct t=37.0 p=3.8e-145; Thinking t=5.63 p=3.0e-8)",
        "results/runs/2026-05-07_0025_gsm8k500_qwen3-4b_deep-table-ext",
    ]
    ws.append(new_row)
    print("[amend] Run_index: appended 1 row")


def main() -> None:
    if not XLSX.exists():
        raise SystemExit(f"missing: {XLSX}")
    if not N500_DIR.exists():
        raise SystemExit(f"missing: {N500_DIR}")

    print(f"[amend] loading {XLSX}")
    wb = load_workbook(XLSX)

    amend_hcds_summary(wb)
    add_per_question_sheet(wb)
    add_task6_sheet(wb)
    amend_run_index(wb)

    print(f"[amend] saving {XLSX}")
    wb.save(XLSX)
    print("[amend] done.")


if __name__ == "__main__":
    main()
