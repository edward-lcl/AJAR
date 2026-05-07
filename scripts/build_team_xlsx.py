#!/usr/bin/env python3
"""Bundle the team-facing CSVs into a single .xlsx with one sheet
per file. Imports cleanly into Google Sheets — drag-and-drop the
.xlsx and Sheets opens it as a multi-tab spreadsheet.

The CSVs are read from
``results/runs/<run_id>/team_spreadsheet/`` (built by
``build_team_spreadsheet.py``). The .xlsx lands in the same
directory.
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_RUN_ID = "2026-05-04_2232_gsm8k50_qwen3-4b_deep-table"

# Sheet name → source CSV name. Sheets ordered for readability when
# the team opens it: highest-signal first.
SHEET_ORDER = [
    ("README",                "README.md"),
    ("HCDS_summary_all",      "hcds_summary_all.csv"),
    ("Per-question_HCDS",     "per_question_hcds.csv"),
    ("Per-question_features", "per_question_features.csv"),
    ("Anchor_sensitivity",    "anchor_sensitivity.csv"),
    ("Length-matched_HCDS",   "length_matched_hcds.csv"),
    ("Wide_baseline_n=500",   "baseline_wide_n500.csv"),
    ("Run_index",             "run_index.csv"),
]

HEADER_FILL = PatternFill(
    start_color="2166AC", end_color="2166AC", fill_type="solid"
)
HEADER_FONT = Font(color="FFFFFF", bold=True)
BANNER_FILL = PatternFill(
    start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"
)


def coerce(value: str):
    """Convert a CSV-string cell to int/float when it parses cleanly,
    otherwise leave as str. Excel renders numeric columns as numbers
    rather than strings, which is what the team wants."""
    if value == "":
        return None
    try:
        i = int(value)
        # Don't lose leading-zero IDs etc., but those would normally
        # have been quoted in the CSV anyway.
        return i
    except ValueError:
        pass
    try:
        return float(value)
    except ValueError:
        return value


def write_csv_sheet(wb: Workbook, sheet_name: str, csv_path: Path) -> None:
    ws = wb.create_sheet(sheet_name)
    if not csv_path.exists():
        ws["A1"] = f"(file not found: {csv_path.name})"
        return

    with csv_path.open("r", newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        ws["A1"] = "(empty file)"
        return

    header = rows[0]
    for col_idx, col in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for r_idx, row in enumerate(rows[1:], start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=coerce(value))

    # Freeze top row + auto-size columns to a sane cap so the sheet
    # is browsable on first open.
    ws.freeze_panes = "A2"
    for c_idx, col in enumerate(header, start=1):
        max_len = len(col)
        for r in rows[1:]:
            if c_idx - 1 < len(r):
                max_len = max(max_len, len(r[c_idx - 1]))
        ws.column_dimensions[get_column_letter(c_idx)].width = min(
            max(max_len + 2, 10), 60
        )


def write_readme_sheet(wb: Workbook, readme_path: Path) -> None:
    """Render the team_spreadsheet README as a single text column,
    one line per row. Excel doesn't render markdown; this just
    preserves the content so the team can see "what's in here"
    without leaving the spreadsheet."""
    ws = wb.create_sheet("README", 0)  # leftmost
    ws.column_dimensions["A"].width = 100
    title_cell = ws.cell(
        row=1, column=1, value="AJAR — team spreadsheet (auto-generated)"
    )
    title_cell.fill = BANNER_FILL
    title_cell.font = Font(bold=True, size=14)
    if not readme_path.exists():
        ws.cell(row=2, column=1, value="(README not found)")
        return
    text = readme_path.read_text(encoding="utf-8")
    for r_idx, line in enumerate(text.splitlines(), start=3):
        cell = ws.cell(row=r_idx, column=1, value=line)
        cell.alignment = Alignment(wrap_text=False, vertical="top")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--run-id", default=DEFAULT_RUN_ID)
    ap.add_argument(
        "--source-subdir",
        default="team_spreadsheet",
        help="Subdirectory under the run dir holding the source CSVs.",
    )
    ap.add_argument(
        "--out",
        type=Path,
        default=None,
        help="Output .xlsx path (default: <run_dir>/<source_subdir>/AJAR_team.xlsx)",
    )
    args = ap.parse_args()

    src_dir = (
        REPO_ROOT / "results" / "runs" / args.run_id / args.source_subdir
    )
    if not src_dir.exists():
        print(
            f"error: source dir not found: {src_dir}.\n"
            f"  run `python3 scripts/build_team_spreadsheet.py` first."
        )
        return 2

    out_path = args.out or (src_dir / "AJAR_team.xlsx")
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Drop the default Sheet so our ordered sheets are clean.
    default = wb.active
    wb.remove(default)

    for sheet_name, fname in SHEET_ORDER:
        if fname == "README.md":
            write_readme_sheet(wb, src_dir / fname)
        else:
            write_csv_sheet(wb, sheet_name, src_dir / fname)

    wb.save(out_path)
    print(f"wrote {out_path.relative_to(REPO_ROOT)}")
    print(f"  sheets: {[s for s, _ in SHEET_ORDER]}")
    return 0


if __name__ == "__main__":
    import sys

    sys.exit(main())
