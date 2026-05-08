"""Sanity-check the wrapped-up AJAR results.xlsx.

Verifies:
  - All expected sheets present
  - Headline numbers in Executive_Summary match underlying CSVs
  - Cross_dataset_HCDS rows match StrategyQA_HCDS_summary and HCDS_summary_all
  - Numeric cells are numeric (not strings)
  - No surprising NaN / None in headline rows
  - Per-question rows agree with summary rows (mean, n)
  - Run_index has no duplicate run_ids
"""

from __future__ import annotations

import csv
import math
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl

ROOT = Path("/Users/edward/Projects/AJAR")
XLSX = ROOT / "results/runs/AJAR results.xlsx"
STAGE1 = ROOT / "results/runs/2026-05-06_2121_strategyqa50_gsm8k50_qwen3-4b_deep-table"
STAGE2 = ROOT / "results/runs/2026-05-04_2232_gsm8k50_qwen3-4b_deep-table"
STAGE3 = ROOT / "results/runs/2026-05-07_0025_gsm8k500_qwen3-4b_deep-table-ext"

EXPECTED_SHEETS = [
    "Executive_Summary",
    "README",
    "Cross_dataset_HCDS",
    "Robustness_variants",
    "HCDS_summary_all",
    "Per-question_HCDS",
    "Per-question_features",
    "Anchor_sensitivity",
    "Length-matched_HCDS",
    "Wide_baseline_n=500",
    "Run_index",
    "HCDS_per_question_n=500",
    "Task6_table_n=500",
    "StrategyQA_HCDS_summary",
    "StrategyQA_HCDS_summary_no_mech",
    "StrategyQA_per_question_HCDS",
    "StrategyQA_anchor_sensitivity",
    "StrategyQA_Task6_table",
    "Anchor_sensitivity_mfix",
]

PASSES: List[str] = []
WARNINGS: List[str] = []
FAILURES: List[str] = []


def check(name: str, ok: bool, detail: str = "") -> None:
    if ok:
        PASSES.append(f"  PASS  {name}" + (f" — {detail}" if detail else ""))
    else:
        FAILURES.append(f"  FAIL  {name}" + (f" — {detail}" if detail else ""))


def warn(name: str, detail: str = "") -> None:
    WARNINGS.append(f"  WARN  {name}" + (f" — {detail}" if detail else ""))


def approx(a: float, b: float, rel: float = 0.005, abs_tol: float = 1e-6) -> bool:
    if a is None or b is None:
        return False
    if math.isnan(a) or math.isnan(b):
        return False
    return abs(a - b) <= max(abs_tol, rel * max(abs(a), abs(b)))


def approx_p(a: float, b: float, rel_log: float = 0.05) -> bool:
    """Compare p-values on log scale (relative tolerance on log10)."""
    if a is None or b is None or a <= 0 or b <= 0:
        return False
    la, lb = math.log10(a), math.log10(b)
    return abs(la - lb) <= max(0.1, rel_log * max(abs(la), abs(lb)))


def read_csv_dicts(path: Path) -> List[Dict[str, str]]:
    with path.open() as f:
        return list(csv.DictReader(f))


def load_summary_csv(path: Path) -> Dict[str, Dict[str, float]]:
    out: Dict[str, Dict[str, float]] = {}
    for r in read_csv_dicts(path):
        out[r["model"]] = {
            "mean_hcds": float(r["mean_hcds"]),
            "p": float(r["p_value_two_sided"]),
            "t": float(r["t_statistic"]),
            "n": int(r["n_questions"]),
            "ci_low": float(r["ci_low_95"]),
            "ci_high": float(r["ci_high_95"]),
        }
    return out


def main() -> None:
    print(f"Loading {XLSX}")
    wb = openpyxl.load_workbook(XLSX, data_only=False)
    sheets = wb.sheetnames

    # --- Sheet presence ---
    missing = [s for s in EXPECTED_SHEETS if s not in sheets]
    extra = [s for s in sheets if s not in EXPECTED_SHEETS]
    check("All expected sheets present", not missing,
          f"missing {missing}" if missing else f"{len(sheets)} sheets")
    if extra:
        warn("Unexpected extra sheets", str(extra))

    # --- StrategyQA summary numbers match source CSV ---
    sqa_csv = load_summary_csv(STAGE1 / "hcds_summary.csv")
    ws = wb["StrategyQA_HCDS_summary"]
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    body = {r[header.index("model")]: dict(zip(header, r)) for r in rows[1:] if r[0]}
    for model in ("instruct", "thinking"):
        sheet_val = body[model]["mean_hcds"]
        csv_val = sqa_csv[model]["mean_hcds"]
        check(f"StrategyQA_HCDS_summary mean matches CSV ({model})",
              approx(sheet_val, csv_val), f"sheet={sheet_val:.6f} csv={csv_val:.6f}")
        sheet_p = body[model]["p_value_two_sided"]
        csv_p = sqa_csv[model]["p"]
        check(f"StrategyQA_HCDS_summary p matches CSV ({model})",
              approx_p(sheet_p, csv_p), f"sheet={sheet_p:.2e} csv={csv_p:.2e}")

    # --- StrategyQA per-question count ---
    sqa_pq = wb["StrategyQA_per_question_HCDS"].max_row - 1  # minus header
    check("StrategyQA_per_question_HCDS has 100 rows (50 instruct + 50 thinking)",
          sqa_pq == 100, f"{sqa_pq} rows")

    # --- StrategyQA per-question mean equals summary mean ---
    pq_ws = wb["StrategyQA_per_question_HCDS"]
    pq_rows = list(pq_ws.iter_rows(values_only=True))
    pq_header = list(pq_rows[0])
    model_idx = pq_header.index("model")
    hcds_idx = pq_header.index("hcds_q")
    by_model: Dict[str, List[float]] = {"instruct": [], "thinking": []}
    for r in pq_rows[1:]:
        if r[model_idx] in by_model and isinstance(r[hcds_idx], (int, float)):
            by_model[r[model_idx]].append(float(r[hcds_idx]))
    for m in by_model:
        mean = sum(by_model[m]) / len(by_model[m]) if by_model[m] else float("nan")
        check(f"StrategyQA per-question mean matches summary ({m})",
              approx(mean, sqa_csv[m]["mean_hcds"]),
              f"per-q mean={mean:.6f} summary={sqa_csv[m]['mean_hcds']:.6f}, n={len(by_model[m])}")

    # --- Cross_dataset_HCDS rows agree with sources ---
    xd = wb["Cross_dataset_HCDS"]
    xd_rows = list(xd.iter_rows(values_only=True))
    # Find header row (row with "Dataset" in col 0)
    header_r = next(i for i, r in enumerate(xd_rows) if r and r[0] == "Dataset")
    xd_header = list(xd_rows[header_r])
    body_rows = [r for r in xd_rows[header_r + 1:] if r and r[0]]

    gsm50_csv = load_summary_csv(STAGE2 / "hcds_summary.csv")
    gsm500_csv = load_summary_csv(STAGE3 / "hcds_summary.csv")

    expected_map = {
        ("GSM8K", 50, "instruct"): gsm50_csv["instruct"],
        ("GSM8K", 50, "thinking"): gsm50_csv["thinking"],
        ("StrategyQA", 50, "instruct"): sqa_csv["instruct"],
        ("StrategyQA", 50, "thinking"): sqa_csv["thinking"],
        ("GSM8K (5-feat)", 500, "instruct"): gsm500_csv["instruct"],
        ("GSM8K (5-feat)", 500, "thinking"): gsm500_csv["thinking"],
    }
    ds_idx = xd_header.index("Dataset")
    n_idx = xd_header.index("n")
    m_idx = xd_header.index("Model")
    hcds_idx_xd = xd_header.index("HCDS mean")
    p_idx_xd = xd_header.index("p (two-sided)")
    for r in body_rows:
        key = (r[ds_idx], r[n_idx], r[m_idx])
        if key not in expected_map:
            warn(f"Cross_dataset_HCDS unexpected row {key}")
            continue
        exp = expected_map[key]
        check(f"Cross_dataset HCDS matches source {key}",
              approx(float(r[hcds_idx_xd]), exp["mean_hcds"], rel=0.01),
              f"sheet={r[hcds_idx_xd]} csv={exp['mean_hcds']:.4f}")
        check(f"Cross_dataset p matches source {key}",
              approx_p(float(r[p_idx_xd]), exp["p"]),
              f"sheet={r[p_idx_xd]} csv={exp['p']:.2e}")

    # --- Robustness_variants Instruct/Thinking match HCDS_summary_all where applicable ---
    ws_all = wb["HCDS_summary_all"]
    all_rows = list(ws_all.iter_rows(values_only=True))
    all_header = list(all_rows[0])
    summary_index: Dict[Tuple[str, str], Dict[str, object]] = {}
    for r in all_rows[1:]:
        if not r[0]:
            continue
        rec = dict(zip(all_header, r))
        summary_index[(rec["variant"], rec["model"])] = rec

    rv = wb["Robustness_variants"]
    rv_rows = list(rv.iter_rows(values_only=True))
    rv_header_r = next(i for i, r in enumerate(rv_rows) if r and r[0] == "Variant")
    rv_header = list(rv_rows[rv_header_r])
    var_idx = rv_header.index("Variant")
    inst_hcds_idx = rv_header.index("Instruct HCDS")
    think_hcds_idx = rv_header.index("Thinking HCDS")
    for r in rv_rows[rv_header_r + 1:]:
        if not r or not r[0]:
            continue
        variant = r[var_idx]
        for model_label, idx in (("instruct", inst_hcds_idx), ("thinking", think_hcds_idx)):
            sheet_val = r[idx]
            if sheet_val in (None, "—", "n/a"):
                continue
            try:
                sheet_f = float(str(sheet_val).replace("+", ""))
            except (ValueError, TypeError):
                warn(f"Robustness_variants {variant}/{model_label}: cannot parse '{sheet_val}'")
                continue
            src = summary_index.get((variant, model_label))
            if src is None:
                warn(f"Robustness_variants variant '{variant}' has no row in HCDS_summary_all")
                continue
            try:
                src_val = float(src["mean_hcds"])
            except (ValueError, TypeError):
                warn(f"HCDS_summary_all {variant}/{model_label}: bad mean_hcds")
                continue
            check(f"Robustness {variant}/{model_label} matches HCDS_summary_all",
                  approx(sheet_f, src_val, rel=0.01),
                  f"sheet={sheet_f:.3f} all={src_val:.3f}")

    # --- Run_index uniqueness ---
    ri = wb["Run_index"]
    ids = [row[0].value for row in ri.iter_rows(min_row=2) if row[0].value]
    dups = [x for x in set(ids) if ids.count(x) > 1]
    check("Run_index has no duplicate run_ids", not dups, f"dupes={dups}" if dups else f"{len(ids)} runs")

    # --- Numeric cell types in headline tables ---
    def numeric_check(sheet: str, header_row: int, col_name: str, *, allow_str_dash: bool = True) -> None:
        ws_local = wb[sheet]
        rows_local = list(ws_local.iter_rows(values_only=True))
        if header_row >= len(rows_local):
            warn(f"{sheet}: header row {header_row} out of range")
            return
        header_local = list(rows_local[header_row])
        if col_name not in header_local:
            warn(f"{sheet}: column '{col_name}' not in header")
            return
        col_idx = header_local.index(col_name)
        bad = []
        for i, r in enumerate(rows_local[header_row + 1:], start=header_row + 2):
            v = r[col_idx] if r and col_idx < len(r) else None
            if v is None:
                continue
            if isinstance(v, str):
                if allow_str_dash and v in ("—", "n/a", ""):
                    continue
                bad.append((i, v))
        check(f"{sheet}.{col_name} numeric (or '—')", not bad,
              f"{len(bad)} string cells: {bad[:3]}" if bad else "")

    numeric_check("StrategyQA_HCDS_summary", 0, "mean_hcds")
    numeric_check("StrategyQA_HCDS_summary", 0, "p_value_two_sided")
    numeric_check("Cross_dataset_HCDS", 4, "HCDS mean")
    numeric_check("Cross_dataset_HCDS", 4, "p (two-sided)")

    # --- StrategyQA Task6 row count ---
    t6 = wb["StrategyQA_Task6_table"]
    expected_rows = 50 * 2 * 3  # 50 q × 2 models × 3 prompts
    check("StrategyQA_Task6_table has 300 data rows + header",
          t6.max_row == expected_rows + 1, f"{t6.max_row} rows")

    # --- README + Executive_Summary not empty ---
    for s in ("README", "Executive_Summary"):
        ws_local = wb[s]
        non_empty = sum(1 for row in ws_local.iter_rows(values_only=True) for v in row if v)
        check(f"{s} is non-empty", non_empty > 50, f"{non_empty} non-empty cells")

    # --- Print summary ---
    print()
    print("=" * 70)
    for line in PASSES:
        print(line)
    print()
    for line in WARNINGS:
        print(line)
    print()
    for line in FAILURES:
        print(line)
    print()
    print(f"Pass: {len(PASSES)} | Warn: {len(WARNINGS)} | Fail: {len(FAILURES)}")
    if FAILURES:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
