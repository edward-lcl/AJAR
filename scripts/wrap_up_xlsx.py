"""One-shot wrap-up: append cross-dataset sheets, executive summary, and README.

Reads CSVs from the three result run directories and writes new sheets into
results/runs/AJAR results.xlsx. Existing sheets are left intact except where
explicitly noted (Run_index gets a new row; README and Executive_Summary are
overwritten because they were placeholders).
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path("/Users/edward/Projects/AJAR")
XLSX = ROOT / "results/runs/AJAR results.xlsx"

STAGE1 = ROOT / "results/runs/2026-05-06_2121_strategyqa50_gsm8k50_qwen3-4b_deep-table"
STAGE2 = ROOT / "results/runs/2026-05-04_2232_gsm8k50_qwen3-4b_deep-table"
STAGE3 = ROOT / "results/runs/2026-05-07_0025_gsm8k500_qwen3-4b_deep-table-ext"

HEADER_FILL = PatternFill("solid", fgColor="4F81BD")
HEADER_FONT = Font(bold=True, color="FFFFFF")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
GOOD_FILL = PatternFill("solid", fgColor="E2EFDA")
BAD_FILL = PatternFill("solid", fgColor="FCE4D6")
TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=11)


def read_csv(path: Path) -> List[List[str]]:
    with path.open() as f:
        return [row for row in csv.reader(f)]


def coerce(value: str):
    if value is None or value == "":
        return None
    if value.startswith("<"):
        return value
    try:
        if "." in value or "e" in value.lower():
            return float(value)
        return int(value)
    except (ValueError, AttributeError):
        return value


def write_csv_to_sheet(wb, sheet_name: str, csv_path: Path, *, replace: bool = True) -> None:
    if sheet_name in wb.sheetnames:
        if replace:
            del wb[sheet_name]
        else:
            return
    ws = wb.create_sheet(sheet_name)
    rows = read_csv(csv_path)
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx,
                           value=coerce(value) if r_idx > 1 else value)
            if r_idx == 1:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
    autosize(ws)


def autosize(ws, max_width: int = 60) -> None:
    for col in ws.columns:
        col = list(col)
        if not col:
            continue
        letter = get_column_letter(col[0].column)
        widest = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[letter].width = min(max(widest + 2, 10), max_width)


def write_grid(ws, start_row: int, start_col: int, grid: List[List[object]],
               *, header: bool = True) -> int:
    for r_idx, row in enumerate(grid):
        for c_idx, value in enumerate(row):
            cell = ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=value)
            if header and r_idx == 0:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
    return start_row + len(grid)


def add_strategyqa_sheets(wb) -> None:
    write_csv_to_sheet(wb, "StrategyQA_HCDS_summary", STAGE1 / "hcds_summary.csv")
    write_csv_to_sheet(wb, "StrategyQA_HCDS_summary_no_mech", STAGE1 / "hcds_summary_no_mech.csv")
    write_csv_to_sheet(wb, "StrategyQA_per_question_HCDS", STAGE1 / "hcds_per_question.csv")
    write_csv_to_sheet(wb, "StrategyQA_anchor_sensitivity", STAGE1 / "task10_anchor_sensitivity.csv")
    write_csv_to_sheet(wb, "StrategyQA_Task6_table", STAGE1 / "task6_table.csv")


def add_anchor_mfix_sheet(wb) -> None:
    write_csv_to_sheet(wb, "Anchor_sensitivity_mfix",
                       STAGE2 / "task10_anchor_sensitivity_mfix_alpha0.5.csv")
    ws = wb["Anchor_sensitivity_mfix"]
    ws.cell(row=ws.max_row + 2, column=1,
            value="Methodology-fix variant of anchor sensitivity (alpha=0.5).")
    ws.cell(row=ws.max_row + 1, column=1,
            value="Re-runs Thinking + explicit_cot only with adjusted methodology to test sensitivity.")
    ws.cell(row=ws.max_row + 1, column=1,
            value=("Compare to Anchor_sensitivity sheet: original Thinking/explicit_cot AmCD = -0.275; "
                   "mfix AmCD = -0.225. Direction unchanged, magnitude similar — methodology choice "
                   "is not driving the negative anchor effect."))


def add_cross_dataset_sheet(wb) -> None:
    if "Cross_dataset_HCDS" in wb.sheetnames:
        del wb["Cross_dataset_HCDS"]
    ws = wb.create_sheet("Cross_dataset_HCDS", index=2)

    ws.cell(row=1, column=1, value="Cross-dataset HCDS replication (Qwen3-4B)").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    ws.cell(row=3, column=1,
            value="HCDS > 0 ⇒ Neutral prompt behaves more like CoT than No-CoT (i.e., model reasons even without being told to).").font = SECTION_FONT
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=8)

    grid: List[List[object]] = [
        ["Dataset", "n", "Model", "HCDS mean", "95% CI", "t", "p (two-sided)", "Verdict"],
        ["GSM8K", 50, "instruct", 2.254, "[1.687, 2.846]", 7.96, 2.21e-10, "neutral aligns with CoT"],
        ["GSM8K", 50, "thinking", 0.479, "[0.068, 0.859]", 2.39, 2.06e-2, "neutral aligns with CoT"],
        ["StrategyQA", 50, "instruct", 1.871, "[1.482, 2.282]", 9.15, 3.53e-12, "neutral aligns with CoT"],
        ["StrategyQA", 50, "thinking", 0.597, "[0.242, 0.950]", 3.24, 2.15e-3, "neutral aligns with CoT"],
        ["GSM8K (5-feat)", 500, "instruct", 2.375, "[2.247, 2.500]", 36.14, 1.98e-141, "neutral aligns with CoT"],
        ["GSM8K (5-feat)", 500, "thinking", 0.330, "[0.206, 0.447]", 5.29, 1.85e-7, "neutral aligns with CoT"],
    ]
    end_row = write_grid(ws, 5, 1, grid)

    ws.cell(row=end_row + 2, column=1, value="Headline takeaways").font = SECTION_FONT
    notes = [
        "1. Direction holds across both datasets, both models, and at both scales (n=50, n=500).",
        "2. Instruct shows a stronger effect than Thinking (~4× larger HCDS) on both datasets — the Thinking model's reasoning is less prompt-conditional.",
        "3. n=500 GSM8K Instruct p=1.98e-141 is overwhelming; n=500 Thinking p=1.85e-7 still highly significant.",
        "4. StrategyQA replication uses 6-feature HCDS (full pipeline incl. mechanistic) on n=50; same direction, p<0.01 both models.",
        "5. The 5-feature (no_mech) variant on StrategyQA gives Instruct +2.57 (p=4.7e-17) and Thinking +0.61 (p=7.6e-4) — robust to feature choice.",
    ]
    for i, note in enumerate(notes, start=end_row + 3):
        ws.cell(row=i, column=1, value=note)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)

    autosize(ws, max_width=80)


def add_robustness_sheet(wb) -> None:
    if "Robustness_variants" in wb.sheetnames:
        del wb["Robustness_variants"]
    ws = wb.create_sheet("Robustness_variants", index=3)

    ws.cell(row=1, column=1, value="HCDS robustness across feature subsets (GSM8K n=50)").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    ws.cell(row=3, column=1,
            value=("Each row drops a different feature (or family) from the HCDS vector and recomputes. "
                   "Goal: verify the signal isn't being driven by any single feature.")).font = SECTION_FONT
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=10)

    grid: List[List[object]] = [
        ["Variant", "Features dropped", "Instruct HCDS", "Instruct p", "Thinking HCDS", "Thinking p", "Verdict"],
    ]
    variants = [
        ("full", []),
        ("no_paraphrase", ["paraphrase_consistency"]),
        ("no_perturb", ["perturbation_delta_accuracy"]),
        ("no_entropy", ["token_entropy_mean", "token_entropy_slope"]),
        ("no_mech", ["mechanistic_intervention_delta_accuracy"]),
        ("no_para_no_mech", ["paraphrase_consistency", "mechanistic_intervention_delta_accuracy"]),
        ("entropy_only", ["latency_per_output_token", "paraphrase_consistency",
                          "perturbation_delta_accuracy", "mechanistic_intervention_delta_accuracy"]),
        ("latency_only", ["token_entropy_mean", "token_entropy_slope", "paraphrase_consistency",
                          "perturbation_delta_accuracy", "mechanistic_intervention_delta_accuracy"]),
    ]

    for variant, dropped in variants:
        if variant == "full":
            csv_path = STAGE2 / "hcds_summary.csv"
        else:
            csv_path = STAGE2 / f"hcds_summary_{variant}.csv"
        if not csv_path.exists():
            grid.append([variant, ", ".join(dropped) or "—", "n/a", "n/a", "n/a", "n/a", "FILE MISSING"])
            continue
        rows = read_csv(csv_path)
        if len(rows) < 3:
            grid.append([variant, ", ".join(dropped) or "—", "n/a", "n/a", "n/a", "n/a", "EMPTY"])
            continue
        idx = {h: i for i, h in enumerate(rows[0])}
        body = {r[idx["model"]]: r for r in rows[1:]}
        i = body.get("instruct", [None] * len(rows[0]))
        t = body.get("thinking", [None] * len(rows[0]))

        def fmt(v, *, sci: bool = False) -> str:
            if v is None or v == "":
                return "—"
            try:
                f = float(v)
            except ValueError:
                return v
            return f"{f:.2e}" if sci else f"{f:+.3f}"

        grid.append([
            variant,
            ", ".join(dropped) or "—",
            fmt(i[idx["mean_hcds"]]) if i[idx["mean_hcds"]] is not None else "—",
            fmt(i[idx["p_value_two_sided"]], sci=True) if i[idx["p_value_two_sided"]] is not None else "—",
            fmt(t[idx["mean_hcds"]]) if t[idx["mean_hcds"]] is not None else "—",
            fmt(t[idx["p_value_two_sided"]], sci=True) if t[idx["p_value_two_sided"]] is not None else "—",
            "robust" if (i[idx["verdict"]] == t[idx["verdict"]] == "neutral aligns with CoT (HCDS > 0, p < 0.05)") else "mixed",
        ])

    end_row = write_grid(ws, 5, 1, grid)

    ws.cell(row=end_row + 2, column=1, value="Conclusion").font = SECTION_FONT
    ws.cell(row=end_row + 3, column=1,
            value=("Signal is largely robust: 7/8 variants show HCDS > 0, p < 0.05 for both models. "
                   "Exception: dropping entropy features kills Thinking-model significance (+0.09, p=0.62) "
                   "while leaving Instruct intact (+1.72, p=7.4e-9). Interpretation: token-level entropy is "
                   "load-bearing for detecting Thinking's hidden reasoning, but not for Instruct's. "
                   "Reported transparently."))
    ws.merge_cells(start_row=end_row + 3, start_column=1, end_row=end_row + 3, end_column=10)
    autosize(ws, max_width=80)


def add_executive_summary(wb) -> None:
    if "Executive_Summary" in wb.sheetnames:
        del wb["Executive_Summary"]
    ws = wb.create_sheet("Executive_Summary", index=0)

    ws.cell(row=1, column=1, value="AJAR — Executive Summary").font = Font(bold=True, size=18)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    ws.cell(row=2, column=1,
            value="Are LLMs Just Acting Reasonable? Hidden CoT Detection in Qwen3-4B (Instruct vs Thinking)").font = Font(italic=True, size=11)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    headline = [
        ("HEADLINE", "Both Qwen3-4B variants reason internally even when prompted not to."),
        ("", "HCDS > 0 (p < 0.05) on every dataset × model tested, with one feature-ablation caveat (see Robustness)."),
        ("", ""),
        ("KEY EVIDENCE", ""),
        ("  GSM8K n=500 Instruct", "HCDS = +2.38, t = 36.14, p = 1.98e-141"),
        ("  GSM8K n=500 Thinking", "HCDS = +0.33, t = 5.29, p = 1.85e-7"),
        ("  GSM8K n=50 Instruct", "HCDS = +2.25, p = 2.2e-10 (full 6-feature pipeline)"),
        ("  GSM8K n=50 Thinking", "HCDS = +0.48, p = 2.1e-2 (full 6-feature pipeline)"),
        ("  StrategyQA n=50 Instruct", "HCDS = +1.87, p = 3.5e-12  (cross-dataset replication)"),
        ("  StrategyQA n=50 Thinking", "HCDS = +0.60, p = 2.1e-3  (cross-dataset replication)"),
        ("", ""),
        ("ROBUSTNESS", ""),
        ("  Feature ablations", "8 variants tested. 7/8 show HCDS > 0, p < 0.05 on both models. Exception: dropping entropy features (no_entropy) loses Thinking-model significance (HCDS +0.09, p=0.62) — entropy is load-bearing for the Thinking signal but not for Instruct (which stays at +1.72, p=7.4e-9). See Robustness_variants."),
        ("  Length control", "Holds for short and long outputs on Instruct. On Thinking, signal is weaker for medium-length outputs (see Length-matched_HCDS)."),
        ("  Methodology variant", "Anchor sensitivity rerun with alpha=0.5 (mfix) gives same direction & similar magnitude as primary methodology."),
        ("", ""),
        ("MECHANISTIC", ""),
        ("  Anchor intervention (Instruct)", "Anchor disruption causes larger accuracy drop than control on neutral_strict; supports localized causal structure."),
        ("  Anchor intervention (Thinking)", "Anchor sensitivity NEGATIVE on Thinking + explicit_cot — long reasoning chains distribute causal load across many steps. Reported transparently as a finding, not a bug."),
        ("", ""),
        ("INTERPRETATION", ""),
        ("  Instruct model", "Strong, prompt-conditional reasoning. CoT instruction matters less than expected — neutral prompt produces near-CoT behavior."),
        ("  Thinking model", "Weaker but still significant HCDS. Reasoning is less prompt-conditional and more architecturally-baked-in. Cannot suppress reasoning even when explicitly asked (explicit_no_cot still produces ~575 reasoning tokens)."),
        ("  Implication", "Behavioral 'no-CoT' baselines are confounded for reasoning-tuned models; HCDS provides a cleaner detection signal."),
        ("", ""),
        ("CAVEATS (transparent reporting)", ""),
        ("  Single seed", "AJAR_SEED=17 throughout. No variance-across-seeds estimate."),
        ("  Single trial per condition", "Deterministic decoding; no temperature sampling distribution."),
        ("  GSM8K contamination", "Not directly mitigated. StrategyQA replication addresses generalization beyond GSM8K."),
        ("  Token cap", "Thinking + explicit_cot hits 1024-token cap on ~50% of n=500 outputs (see Wide_baseline_n=500); accuracy may be understated."),
        ("  Mechanistic n", "Anchor sensitivity reported on n=50 only; mechanistic feature in HCDS is conditional on availability."),
        ("", ""),
        ("DELIVERABLES IN THIS WORKBOOK", ""),
        ("  Cross_dataset_HCDS", "Side-by-side GSM8K vs StrategyQA evidence."),
        ("  Robustness_variants", "Feature ablation grid."),
        ("  Wide_baseline_n=500", "Accuracy / latency / output-length at scale."),
        ("  HCDS_summary_all", "9 feature-subset variants on GSM8K n=50."),
        ("  StrategyQA_HCDS_summary", "Cross-dataset replication."),
        ("  Anchor_sensitivity / Anchor_sensitivity_mfix", "Causal validation + methodology robustness."),
        ("  Length-matched_HCDS", "Length-controlled robustness check."),
        ("  Run_index", "Lineage of all six runs feeding this workbook."),
        ("  README", "Provenance, dates, commit hashes, caveats."),
    ]

    for i, (label, value) in enumerate(headline, start=4):
        c1 = ws.cell(row=i, column=1, value=label)
        c2 = ws.cell(row=i, column=2, value=value)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=8)
        if label == "HEADLINE":
            c1.font = Font(bold=True, color="C00000", size=12)
            c2.font = Font(bold=True, size=12)
        elif label and not label.startswith(" ") and value == "":
            c1.font = SECTION_FONT
        elif label.startswith("  "):
            c1.font = Font(italic=True)
        c1.alignment = Alignment(vertical="top", wrap_text=True)
        c2.alignment = Alignment(vertical="top", wrap_text=True)

    ws.column_dimensions["A"].width = 36
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 18


def add_readme(wb) -> None:
    if "README" in wb.sheetnames:
        del wb["README"]
    ws = wb.create_sheet("README", index=1)

    ws.cell(row=1, column=1, value="AJAR — Workbook README").font = Font(bold=True, size=18)

    sections = [
        ("Last updated", "2026-05-08 (submission day)"),
        ("Repo", "github.com/edward/AJAR (local: /Users/edward/Projects/AJAR)"),
        ("Latest commit", "7cd7f7b — n=500 HCDS extension"),
        ("", ""),
        ("PIPELINE LINEAGE", ""),
        ("Stage 0 — Wide baseline", "2026-05-04_1856 (GSM8K n=500, neutral_strict prompt). Confirms generation pipeline + baseline accuracy."),
        ("Stage 1 — StrategyQA-50 deep table", "2026-05-06_2121. Full 6-feature HCDS on StrategyQA. Note: dataset column internally labeled 'GSM8K' but data is StrategyQA (yes/no). HCDS computed 2026-05-08."),
        ("Stage 2 — GSM8K-50 deep table (primary)", "2026-05-04_2232. Headline n=50 mechanistic + behavioral analysis. Full 6-feature HCDS + 8 robustness variants."),
        ("Stage 3 — n=500 HCDS extension", "2026-05-07_0025. 5-feature HCDS (no mech) at n=500 GSM8K. Headline statistical evidence."),
        ("Methodology-fix anchor variant", "2026-05-04_2232 + alpha=0.5 rerun. Validates anchor sensitivity finding under alternative methodology."),
        ("", ""),
        ("DATA COLLECTION DATES", ""),
        ("GSM8K n=500 baseline", "2026-05-04 (Stage 0)"),
        ("GSM8K n=50 deep table", "2026-05-04 (Stage 2)"),
        ("GSM8K n=500 5-feature extension", "2026-05-07 (Stage 3)"),
        ("StrategyQA n=50 deep table", "2026-05-06 → 2026-05-08 (Stage 1, two-night run; resumed 19:43 May 7, completed 03:52 May 8)"),
        ("Methodology-fix anchor (alpha=0.5)", "2026-05-08 (Stage 2 mfix; completed 06:04)"),
        ("", ""),
        ("KEY METHODOLOGICAL CHOICES (HCDS)", ""),
        ("Feature vector", "6 dims: latency_per_output_token, token_entropy_mean, token_entropy_slope, paraphrase_consistency, perturbation_delta_accuracy, mechanistic_intervention_delta_accuracy"),
        ("Distance metric", "Euclidean with partial-feature handling (drop missing dims per question)"),
        ("Normalization", "Per-model z-score over all (prompt × question) rows"),
        ("Score formula", "HCDS_q = D(Neutral, NoCoT) − D(Neutral, CoT). Aggregate: mean across questions."),
        ("Stats", "1000-sample bootstrap percentile CI (seed=17) + one-sample two-sided t-test vs 0"),
        ("n=500 variant", "5-feature (drops mechanistic_intervention_delta_accuracy) since mech feature is n=50 only"),
        ("Reference impl", "scripts/compute_hcds.py"),
        ("", ""),
        ("KEY METHODOLOGICAL CHOICES (Mechanistic)", ""),
        ("Anchor identification", "Token-level attention probing; top-K positions by attention mass on candidate-answer tokens"),
        ("Intervention modes", "attention_zero (zero-out attention) and residual_zero (zero-out residual stream) at anchor layers"),
        ("Control intervention", "Same modes applied at random non-anchor steps"),
        ("Anchor sensitivity score", "anchor_drop − control_drop; positive ⇒ anchors are causally privileged"),
        ("Reference impl", "scripts/run_qwen3_gsm8k_mi.py"),
        ("", ""),
        ("KNOWN LIMITATIONS", ""),
        ("Single seed", "AJAR_SEED=17. No variance-across-seeds estimate."),
        ("Single trial per condition", "Deterministic decoding; no sampling distribution."),
        ("GSM8K contamination", "Not directly mitigated. StrategyQA replication provides cross-dataset evidence."),
        ("Thinking 1024-token cap", "Hits cap on ~50% of explicit_cot outputs at n=500 — accuracy may be understated."),
        ("Thinking ignores no-CoT prompt", "explicit_no_cot still produces ~575 reasoning tokens — confounded as a no-CoT baseline. Reported transparently as a finding."),
        ("Anchor sensitivity (Thinking + explicit_cot)", "Negative direction (control disruption > anchor disruption). Long chains distribute causal load. Reported as a finding."),
        ("Stage 1 dataset label", "task6_table.csv has 'dataset=GSM8K' but data is StrategyQA. Mislabeling at collection time, not analysis time. Verified by question content + yes/no answer keys."),
        ("", ""),
        ("FAILURE LOG", ""),
        ("Stage 1 sample 29 / neutral_strict / baseline", "RuntimeError: Probe position 1573 exceeds attention seq_len 1573. Off-by-one in probe-position clipping vs attention-tensor length when input was truncated. 1 failure / 150 baseline runs in StrategyQA. Other 149 baselines + all interventions clean."),
        ("Stage 2 (mfix)", "0 failures across 50 baselines + 300 interventions."),
        ("", ""),
        ("HOW TO READ THIS WORKBOOK", ""),
        ("Start with", "Executive_Summary"),
        ("Headline evidence", "Cross_dataset_HCDS, Wide_baseline_n=500"),
        ("Robustness", "Robustness_variants, Length-matched_HCDS, Anchor_sensitivity_mfix"),
        ("Mechanistic", "Anchor_sensitivity, StrategyQA_anchor_sensitivity"),
        ("Per-question detail (for spot checks)", "HCDS_per_question_n=500, StrategyQA_per_question_HCDS, Per-question_HCDS, Per-question_features"),
        ("Raw feature tables", "Task6_table_n=500, StrategyQA_Task6_table"),
        ("Lineage", "Run_index"),
    ]

    for i, (label, value) in enumerate(sections, start=3):
        c1 = ws.cell(row=i, column=1, value=label)
        c2 = ws.cell(row=i, column=2, value=value)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=8)
        c1.alignment = Alignment(vertical="top", wrap_text=True)
        c2.alignment = Alignment(vertical="top", wrap_text=True)
        if label and not value:
            c1.font = SECTION_FONT
        elif label and label.isupper():
            c1.font = Font(bold=True, size=12)

    ws.column_dimensions["A"].width = 38
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 22


def update_run_index(wb) -> None:
    ws = wb["Run_index"]
    existing_ids = {row[0].value for row in ws.iter_rows(min_row=2) if row[0].value}

    new_strategyqa = "2026-05-06_2121_strategyqa50_gsm8k50_qwen3-4b_deep-table"
    if new_strategyqa not in existing_ids:
        ws.append([
            new_strategyqa,
            "yes",
            "Cross-dataset replication: StrategyQA n=50, full 6-feature HCDS + anchor sensitivity",
            "Headline: HCDS positive both models on StrategyQA (Instruct +1.87 p=3.5e-12; Thinking +0.60 p=2.1e-3)",
            "results/runs/2026-05-06_2121_strategyqa50_gsm8k50_qwen3-4b_deep-table",
        ])

    # Fix stale headline numbers for the n=500 extension run.
    target = "2026-05-07_0025_gsm8k500_qwen3-4b_deep-table-ext"
    for row in ws.iter_rows(min_row=2):
        if row[0].value == target:
            row[3].value = (
                "Headline: HCDS positive both models at n=500 "
                "(Instruct t=36.14 p=1.98e-141; Thinking t=5.29 p=1.85e-7)"
            )


def main() -> None:
    print(f"Loading {XLSX} ...")
    wb = openpyxl.load_workbook(XLSX)
    print(f"Existing sheets: {wb.sheetnames}")

    print("Adding StrategyQA sheets ...")
    add_strategyqa_sheets(wb)

    print("Adding mfix anchor sheet ...")
    add_anchor_mfix_sheet(wb)

    print("Adding cross-dataset summary ...")
    add_cross_dataset_sheet(wb)

    print("Adding robustness variants ...")
    add_robustness_sheet(wb)

    print("Writing executive summary ...")
    add_executive_summary(wb)

    print("Writing README ...")
    add_readme(wb)

    print("Updating Run_index ...")
    update_run_index(wb)

    if "Chart_Data" in wb.sheetnames and wb["Chart_Data"].max_row <= 1:
        print("Removing empty Chart_Data placeholder ...")
        del wb["Chart_Data"]

    wb.save(XLSX)
    print(f"Saved {XLSX}")
    print(f"Final sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
